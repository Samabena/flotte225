"""
Export service — Sprint 7
  US-031  Export fleet data as PDF or Excel (Pro / Business plan only)
"""

from io import BytesIO
from datetime import datetime

from sqlalchemy.orm import Session

from app.models.activity_log import ActivityLog
from app.models.fuel_entry import FuelEntry
from app.models.maintenance import Maintenance
from app.models.user import User
from app.models.vehicle import Vehicle


# ── Helpers ───────────────────────────────────────────────────────────────────


def _fuel_rows(db: Session, owner_id: int):
    """Return (FuelEntry, vehicle_name, driver_name) for an owner, newest first."""
    return (
        db.query(FuelEntry, Vehicle.name.label("vname"), User.full_name.label("dname"))
        .join(Vehicle, FuelEntry.vehicle_id == Vehicle.id)
        .join(User, FuelEntry.driver_id == User.id)
        .filter(Vehicle.owner_id == owner_id)
        .order_by(FuelEntry.date.desc())
        .all()
    )


def _maintenance_rows(db: Session, owner_id: int):
    """Return (Maintenance, vehicle_name) for an owner."""
    return (
        db.query(Maintenance, Vehicle.name.label("vname"))
        .join(Vehicle, Maintenance.vehicle_id == Vehicle.id)
        .filter(Vehicle.owner_id == owner_id)
        .order_by(Vehicle.name)
        .all()
    )


# ── Excel ─────────────────────────────────────────────────────────────────────


def generate_fuel_excel(db: Session, owner_id: int) -> BytesIO:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill

    GREEN = "005F02"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Carburant"

    headers = [
        "Date",
        "Véhicule",
        "Chauffeur",
        "Odomètre (km)",
        "Quantité (L)",
        "Montant (FCFA)",
        "Conso. (L/100km)",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color=GREEN, end_color=GREEN, fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    for entry, vname, dname in _fuel_rows(db, owner_id):
        conso = (
            float(entry.consumption_per_100km) if entry.consumption_per_100km else ""
        )
        ws.append(
            [
                entry.date.isoformat(),
                vname,
                dname,
                entry.odometer_km,
                float(entry.quantity_litres),
                float(entry.amount_fcfa),
                conso,
            ]
        )

    for col in ws.columns:
        width = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max(12, min(width + 2, 40))

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def generate_maintenance_excel(db: Session, owner_id: int) -> BytesIO:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill

    GREEN = "005F02"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Maintenance"

    headers = [
        "Véhicule",
        "Dernière vidange (km)",
        "Expiration assurance",
        "Expiration contrôle technique",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color=GREEN, end_color=GREEN, fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    for record, vname in _maintenance_rows(db, owner_id):
        ws.append(
            [
                vname,
                record.last_oil_change_km or "—",
                record.insurance_expiry.isoformat() if record.insurance_expiry else "—",
                (
                    record.inspection_expiry.isoformat()
                    if record.inspection_expiry
                    else "—"
                ),
            ]
        )

    for col in ws.columns:
        width = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max(14, min(width + 2, 40))

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ── PDF ───────────────────────────────────────────────────────────────────────


def _pdf_header_style():
    from reportlab.lib import colors

    return colors.HexColor("#005F02")


def generate_fuel_pdf(db: Session, owner_id: int, owner_name: str) -> BytesIO:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import cm
    from reportlab.platypus import (
        Paragraph,
        SimpleDocTemplate,
        Spacer,
        Table,
        TableStyle,
    )

    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=landscape(A4),
        rightMargin=1.5 * cm,
        leftMargin=1.5 * cm,
        topMargin=2 * cm,
        bottomMargin=2 * cm,
    )
    styles = getSampleStyleSheet()
    GREEN = colors.HexColor("#005F02")

    story = [
        Paragraph(f"Rapport carburant — {owner_name}", styles["Title"]),
        Paragraph(
            f"Généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')}",
            styles["Normal"],
        ),
        Spacer(1, 0.4 * cm),
    ]

    data = [
        [
            "Date",
            "Véhicule",
            "Chauffeur",
            "Odomètre",
            "Qté (L)",
            "Montant (FCFA)",
            "L/100km",
        ]
    ]
    for entry, vname, dname in _fuel_rows(db, owner_id):
        conso = (
            f"{float(entry.consumption_per_100km):.2f}"
            if entry.consumption_per_100km
            else "—"
        )
        data.append(
            [
                entry.date.strftime("%d/%m/%Y"),
                vname,
                dname,
                f"{entry.odometer_km:,}",
                f"{float(entry.quantity_litres):.2f}",
                f"{float(entry.amount_fcfa):,.0f}",
                conso,
            ]
        )

    if len(data) == 1:
        story.append(Paragraph("Aucune entrée carburant.", styles["Normal"]))
    else:
        table = Table(data, repeatRows=1)
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), GREEN),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, -1), 9),
                    ("GRID", (0, 0), (-1, -1), 0.4, colors.lightgrey),
                    (
                        "ROWBACKGROUNDS",
                        (0, 1),
                        (-1, -1),
                        [colors.white, colors.HexColor("#f5f5f5")],
                    ),
                    ("ALIGN", (3, 1), (-1, -1), "RIGHT"),
                    ("TOPPADDING", (0, 0), (-1, -1), 4),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                ]
            )
        )
        story.append(table)

    doc.build(story)
    buf.seek(0)
    return buf


def generate_maintenance_pdf(db: Session, owner_id: int, owner_name: str) -> BytesIO:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import cm
    from reportlab.platypus import (
        Paragraph,
        SimpleDocTemplate,
        Spacer,
        Table,
        TableStyle,
    )

    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=A4,
        rightMargin=2 * cm,
        leftMargin=2 * cm,
        topMargin=2 * cm,
        bottomMargin=2 * cm,
    )
    styles = getSampleStyleSheet()
    GREEN = colors.HexColor("#005F02")

    story = [
        Paragraph(f"Rapport maintenance — {owner_name}", styles["Title"]),
        Paragraph(
            f"Généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')}",
            styles["Normal"],
        ),
        Spacer(1, 0.4 * cm),
    ]

    data = [["Véhicule", "Dernière vidange (km)", "Expir. assurance", "Expir. CT"]]
    for record, vname in _maintenance_rows(db, owner_id):
        data.append(
            [
                vname,
                str(record.last_oil_change_km) if record.last_oil_change_km else "—",
                (
                    record.insurance_expiry.strftime("%d/%m/%Y")
                    if record.insurance_expiry
                    else "—"
                ),
                (
                    record.inspection_expiry.strftime("%d/%m/%Y")
                    if record.inspection_expiry
                    else "—"
                ),
            ]
        )

    if len(data) == 1:
        story.append(
            Paragraph("Aucun enregistrement de maintenance.", styles["Normal"])
        )
    else:
        table = Table(
            data, repeatRows=1, colWidths=[6 * cm, 4.5 * cm, 4.5 * cm, 4.5 * cm]
        )
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), GREEN),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, -1), 10),
                    ("GRID", (0, 0), (-1, -1), 0.4, colors.lightgrey),
                    (
                        "ROWBACKGROUNDS",
                        (0, 1),
                        (-1, -1),
                        [colors.white, colors.HexColor("#f5f5f5")],
                    ),
                    ("TOPPADDING", (0, 0), (-1, -1), 5),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
                ]
            )
        )
        story.append(table)

    doc.build(story)
    buf.seek(0)
    return buf


# ── Analytics helpers ─────────────────────────────────────────────────────────


def _analytics_data(db: Session, owner_id: int):
    from app.services.dashboard_service import (
        _get_financial_summary,
        _get_consumption_indicators,
    )

    financial = _get_financial_summary(db, owner_id)
    consumption = _get_consumption_indicators(db, owner_id)
    return financial, consumption


def _activity_log_rows(db: Session, owner_id: int):
    """Return activity log rows for an owner, newest first (max 500)."""
    return (
        db.query(
            ActivityLog, User.full_name.label("dname"), Vehicle.name.label("vname")
        )
        .outerjoin(User, ActivityLog.driver_id == User.id)
        .outerjoin(Vehicle, ActivityLog.vehicle_id == Vehicle.id)
        .filter(ActivityLog.owner_id == owner_id)
        .order_by(ActivityLog.created_at.desc())
        .limit(500)
        .all()
    )


# ── Analytics Excel ───────────────────────────────────────────────────────────


def generate_analytics_excel(db: Session, owner_id: int) -> BytesIO:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill

    GREEN = "005F02"
    financial, consumption = _analytics_data(db, owner_id)

    wb = openpyxl.Workbook()

    # Sheet 1 — monthly trend
    ws1 = wb.active
    ws1.title = "Tendance mensuelle"
    headers1 = ["Mois", "Dépenses (FCFA)"]
    ws1.append(headers1)
    for cell in ws1[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color=GREEN, end_color=GREEN, fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    for row in financial.monthly_trend:
        ws1.append([row.month, float(row.spend_fcfa)])
    for col in ws1.columns:
        width = max(len(str(cell.value or "")) for cell in col)
        ws1.column_dimensions[col[0].column_letter].width = max(14, min(width + 2, 40))

    # Sheet 2 — per-vehicle spend + consumption
    ws2 = wb.create_sheet("Par véhicule")
    headers2 = [
        "Véhicule",
        "Dépenses totales (FCFA)",
        "Conso. moy. (L/100km)",
        "Nb entrées",
    ]
    ws2.append(headers2)
    for cell in ws2[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color=GREEN, end_color=GREEN, fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    spend_map = {s.vehicle_id: float(s.spend_fcfa) for s in financial.spend_per_vehicle}
    for ind in consumption:
        ws2.append(
            [
                ind.vehicle_name,
                spend_map.get(ind.vehicle_id, 0),
                (
                    float(ind.avg_consumption_per_100km)
                    if ind.avg_consumption_per_100km
                    else "—"
                ),
                ind.entry_count,
            ]
        )
    for col in ws2.columns:
        width = max(len(str(cell.value or "")) for cell in col)
        ws2.column_dimensions[col[0].column_letter].width = max(14, min(width + 2, 40))

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ── Analytics PDF ─────────────────────────────────────────────────────────────


def generate_analytics_pdf(db: Session, owner_id: int, owner_name: str) -> BytesIO:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import cm
    from reportlab.platypus import (
        Paragraph,
        SimpleDocTemplate,
        Spacer,
        Table,
        TableStyle,
    )

    financial, consumption = _analytics_data(db, owner_id)
    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=A4,
        rightMargin=2 * cm,
        leftMargin=2 * cm,
        topMargin=2 * cm,
        bottomMargin=2 * cm,
    )
    styles = getSampleStyleSheet()
    GREEN = colors.HexColor("#005F02")

    story = [
        Paragraph(f"Rapport analytique — {owner_name}", styles["Title"]),
        Paragraph(
            f"Généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')}", styles["Normal"]
        ),
        Spacer(1, 0.5 * cm),
        Paragraph("Tendance mensuelle des dépenses", styles["Heading2"]),
    ]

    trend_data = [["Mois", "Dépenses (FCFA)"]]
    for row in financial.monthly_trend:
        trend_data.append([row.month, f"{float(row.spend_fcfa):,.0f}"])
    if len(trend_data) > 1:
        t1 = Table(trend_data, colWidths=[5 * cm, 5 * cm], repeatRows=1)
        t1.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), GREEN),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, -1), 9),
                    ("GRID", (0, 0), (-1, -1), 0.4, colors.lightgrey),
                    (
                        "ROWBACKGROUNDS",
                        (0, 1),
                        (-1, -1),
                        [colors.white, colors.HexColor("#f5f5f5")],
                    ),
                    ("ALIGN", (1, 1), (1, -1), "RIGHT"),
                    ("TOPPADDING", (0, 0), (-1, -1), 4),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                ]
            )
        )
        story.append(t1)
    else:
        story.append(Paragraph("Aucune donnée de dépenses.", styles["Normal"]))

    story.append(Spacer(1, 0.5 * cm))
    story.append(Paragraph("Indicateurs par véhicule", styles["Heading2"]))

    spend_map = {s.vehicle_id: float(s.spend_fcfa) for s in financial.spend_per_vehicle}
    veh_data = [["Véhicule", "Dépenses (FCFA)", "L/100km moy.", "Nb entrées"]]
    for ind in consumption:
        veh_data.append(
            [
                ind.vehicle_name,
                f"{spend_map.get(ind.vehicle_id, 0):,.0f}",
                (
                    f"{float(ind.avg_consumption_per_100km):.2f}"
                    if ind.avg_consumption_per_100km
                    else "—"
                ),
                str(ind.entry_count),
            ]
        )
    if len(veh_data) > 1:
        t2 = Table(veh_data, repeatRows=1)
        t2.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), GREEN),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, -1), 9),
                    ("GRID", (0, 0), (-1, -1), 0.4, colors.lightgrey),
                    (
                        "ROWBACKGROUNDS",
                        (0, 1),
                        (-1, -1),
                        [colors.white, colors.HexColor("#f5f5f5")],
                    ),
                    ("ALIGN", (1, 1), (-1, -1), "RIGHT"),
                    ("TOPPADDING", (0, 0), (-1, -1), 4),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                ]
            )
        )
        story.append(t2)
    else:
        story.append(Paragraph("Aucun véhicule actif.", styles["Normal"]))

    doc.build(story)
    buf.seek(0)
    return buf


# ── Activity Log Excel ────────────────────────────────────────────────────────


def generate_activity_log_excel(db: Session, owner_id: int) -> BytesIO:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill

    GREEN = "005F02"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Journal d'activité"

    headers = ["Date", "Action", "Chauffeur", "Véhicule", "Détails"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color=GREEN, end_color=GREEN, fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    for log, dname, vname in _activity_log_rows(db, owner_id):
        ws.append(
            [
                log.created_at.strftime("%d/%m/%Y %H:%M") if log.created_at else "—",
                log.action or "—",
                dname or "—",
                vname or "—",
                str(log.data_after or ""),
            ]
        )

    for col in ws.columns:
        width = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max(12, min(width + 2, 50))

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ── Activity Log PDF ──────────────────────────────────────────────────────────


def generate_activity_log_pdf(db: Session, owner_id: int, owner_name: str) -> BytesIO:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import cm
    from reportlab.platypus import (
        Paragraph,
        SimpleDocTemplate,
        Spacer,
        Table,
        TableStyle,
    )

    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=landscape(A4),
        rightMargin=1.5 * cm,
        leftMargin=1.5 * cm,
        topMargin=2 * cm,
        bottomMargin=2 * cm,
    )
    styles = getSampleStyleSheet()
    GREEN = colors.HexColor("#005F02")

    story = [
        Paragraph(f"Journal d'activité — {owner_name}", styles["Title"]),
        Paragraph(
            f"Généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')}", styles["Normal"]
        ),
        Spacer(1, 0.4 * cm),
    ]

    data = [["Date", "Action", "Chauffeur", "Véhicule", "Détails"]]
    for log, dname, vname in _activity_log_rows(db, owner_id):
        details = str(log.data_after or "")[:80]
        data.append(
            [
                log.created_at.strftime("%d/%m/%Y %H:%M") if log.created_at else "—",
                log.action or "—",
                dname or "—",
                vname or "—",
                details,
            ]
        )

    if len(data) == 1:
        story.append(Paragraph("Aucune activité enregistrée.", styles["Normal"]))
    else:
        table = Table(data, repeatRows=1)
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), GREEN),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, -1), 8),
                    ("GRID", (0, 0), (-1, -1), 0.4, colors.lightgrey),
                    (
                        "ROWBACKGROUNDS",
                        (0, 1),
                        (-1, -1),
                        [colors.white, colors.HexColor("#f5f5f5")],
                    ),
                    ("TOPPADDING", (0, 0), (-1, -1), 3),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
                ]
            )
        )
        story.append(table)

    doc.build(story)
    buf.seek(0)
    return buf
